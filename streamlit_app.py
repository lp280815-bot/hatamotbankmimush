# -*- coding: utf-8 -*-
"""
התאמות בנק – 1 עד 12 (גרסת סכומים קשיחה לכלל 3)
- כלל 1: OV/RC 1:1 (תאריך+סכום)
- כלל 2: הוראות קבע (469/515) + 'הוראת קבע ספקים':
    כל השורות בחובה; שורת סיכום 20001 בזכות = סה״כ חובה
    של שורות עם מס’ ספק. שורות בלי מס’ ספק צבועות כתום.
- כלל 3: העברות (485, 'העב' במקבץ-נט') – מסמן רק אם קיים
    צד בנק וגם צד ספרים ושוויי־סכום (במונחי ערך מוחלט).
    אין דרישת התאמת תאריך. אם אין התאמה → גיליון 'פערי סכומים – כלל 3'.
- כלל 4: שיקים ספקים (493) עם טולרנס סכום על התאמת אסמכתאות (Ref1 בנק ↔ Ref2 ספרים).
- כללים 5–10: לפי הלוגיקה שאישרת.
- 11–12: placeholders.
- עיצוב: RTL, A4 לרוחב, Fit-to-width=1, שוליים נוחים.
"""

import io, os, re, json
from datetime import datetime
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.page import PageMargins

# ---------------- UI ----------------
st.set_page_config(page_title="התאמות בנק – 1 עד 12", page_icon="✅", layout="centered")
st.markdown("""
<style>
html, body, [class*="css"] { direction: rtl; text-align: right; }
.block-container { padding-top: 1rem; max-width: 1100px; }
</style>
""", unsafe_allow_html=True)
st.title("התאמות בנק – 1 עד 12")

# ---------------- Constants ----------------
STANDING_CODES = {469, 515}         # כלל 2
OVRC_CODES     = {120, 175}         # כלל 1
TRANSFER_CODE  = 485                # כלל 3
TRANSFER_PHRASE = "העב' במקבץ-נט"
RULE4_CODE     = 493                # כלל 4
RULE4_EPS      = 0.50               # טולרנס כלל 4

# כלל 3 – התאמת סכומים (0.00 = חייב זהות מוחלטת בערך מוחלט)
RULE3_AMOUNT_EPS = 0.00

# כללים 5–10
RULE5_CODES = {453, 472, 473, 124}  # עמלות – חיובי ועד 1000
RULE6_COMPANY = 'פאיימי בע"מ'       # קוד 175, שלילי, פרטים בדיוק
RULE7_CODE = 143; RULE7_PHRASE = "שיקים ממשמרת"
RULE8_CODE = 191; RULE8_PHRASE = "הפק' שיק-שידור"
RULE9_CODE = 205; RULE9_PHRASE = "הפק.שיק במכונה"
RULE10_CODES = {191, 132, 396}

# ---------------- VLOOKUP default mappings ----------------
VK_FILE = "rules_store.json"

DEFAULT_NAME_MAP = {
    "בזק בינלאומי ב": "30006",
    "פרי ירוחם חב'": "34714",
    "סלקום ישראל בע": "30055",
    "בזק-הוראות קבע": "34746",
    "דרך ארץ הייווי": "34602",
    "גלובס פבלישר ע": "30067",
    "פלאפון תקשורת": "30030",
    "מרכז הכוכביות": "30002",
    "ע.אשדוד-מסים": "30056",
    "א.ש.א(בס\"ד)אחז": "30050",
    "או.פי.ג'י(מ.כ)": "30047",
    "רשות האכיפה וה": "67-1",
    "קול ביז מילניו": "30053",
    "פריוריטי סופטו": "30097",
    "אינטרנט רימון": "34636",
    "עו\"דכנית בע\"מ": "30018",
    "עיריית רמת גן": "30065",
    "פז חברת נפט בע": "34811",
    "ישראכרט": "28002",
    "חברת החשמל ליש": "30015",
    "הפניקס ביטוח": "34686",
    "מימון ישיר מקב": "34002",
    "שלמה טפר": "30247",
    "נמרוד תבור עורך-דין": "30038",
    "עיריית בית שמש": "34805",
    "פז קמעונאות וא": "34811",
    "הו\"ק הלו' רבית": "8004",
}

# placeholders 11–12
def rule11_placeholder(df, match_col, code_col, bamt_col, details_col):
    return df[match_col]

def rule12_placeholder(df, match_col, code_col, bamt_col, details_col):
    return df[match_col]

# ---------------- Column maps ----------------
MATCH_COLS = ["מס.התאמה","מס. התאמה","מס התאמה","מספר התאמה","התאמה"]
BANK_CODES = ["קוד פעולת בנק","קוד פעולה","קוד פעולת","Bank Code"]
BANK_AMTS  = ["סכום בדף","סכום דף","סכום בבנק","סכום תנועת בנק","Bank Amount"]
BOOKS_AMTS = ["סכום בספרים","סכום בספר","סכום ספרים","Books Amount"]
REF1S      = ["אסמכתא 1","אסמכתא1","אסמכתא","אסמכתה","Ref1"]
REF2S      = ["אסמכתא 2","אסמכתא2","אסמכתא-2","אסמכתה 2","Ref2"]
DATES      = ["תאריך מאזן","תאריך ערך","תאריך","Date"]
DETAILS    = ["פרטים","תיאור","שם ספק","Details","תאור"]

# aux (עזר להעברות/כלל 3)
AUX_DATE_KEYS = ["תאריך פריקה","תאריך","פריקה"]   # כולל שעה/חותמת אירוע
AUX_AMT_KEYS  = ["אחרי ניכוי","אחרי","סכום"]
AUX_PAYNO_KEYS= ["מס' תשלום","מס תשלום","מספר תשלום"]

# ---------------- Helpers ----------------
def pick_col(df, names):
    for n in names:
        if n in df.columns:
            return n
    for n in names:
        for c in df.columns:
            if isinstance(c, str) and n in c:
                return c
    return None

def to_num(s):
    s = (s.astype(str)
         .str.replace(",", "", regex=False)
         .str.replace("₪", "", regex=False)
         .str.replace("\u200f", "", regex=False)
         .str.replace("\u200e", "", regex=False)
         .str.strip())
    return pd.to_numeric(s, errors="coerce")

def norm_date(series):
    def f(x):
        if pd.isna(x):
            return pd.NaT
        if isinstance(x, (pd.Timestamp, datetime)):
            return pd.Timestamp(x.date())
        return pd.to_datetime(x, dayfirst=True, errors="coerce").normalize()
    return series.apply(f)

def ws_to_df(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(x) if x is not None else "" for x in rows[0]]
    data = [list(r[:len(header)]) for r in rows[1:]]
    return pd.DataFrame(data, columns=header)

def only_digits(s):
    return re.sub(r"\D", "", str(s)).lstrip("0") or "0"

# ---------------- VLOOKUP store ----------------
def vk_load():
    """טוען את rules_store.json וממזג את רשימת ברירת המחדל."""
    if os.path.exists(VK_FILE):
        try:
            with open(VK_FILE, "r", encoding="utf-8") as f:
                store = json.load(f)
        except Exception:
            store = {"name_map": {}, "amount_map": {}}
    else:
        store = {"name_map": {}, "amount_map": {}}

    # ודא שמפתחות קיימים
    store.setdefault("name_map", {})
    store.setdefault("amount_map", {})

    # מיזוג מיפוי ברירת מחדל – לא לדרוס מה שכבר קיים
    for k, v in DEFAULT_NAME_MAP.items():
        store["name_map"].setdefault(k, v)

    # נשמור בחזרה כדי שהקובץ תמיד יכיל גם את ברירת המחדל
    vk_save(store)
    return store

def vk_save(store):
    with open(VK_FILE, "w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False, indent=2)

def import_name_map_from_excel(file, store):
    """
    ייבוא מיפוי ספקים מאקסל:
    עמודה 1 – 'פרטים' (או כותרת דומה),
    עמודה 2 – 'מס' ספק'.
    """
    df = pd.read_excel(file)

    col_det = pick_col(df, DETAILS) or df.columns[0]
    col_sup = pick_col(df, ["מס' ספק", "מס ספק", "מספר ספק", "ספק", "Supplier", "Supplier No"])
    if col_sup is None:
        if len(df.columns) < 2:
            raise ValueError("הקובץ חייב לפחות שתי עמודות: פרטים ומס' ספק.")
        col_sup = df.columns[1]

    df_sub = df[[col_det, col_sup]].dropna(how="all")
    added = 0

    for _, row in df_sub.iterrows():
        name = str(row[col_det]).strip()
        sup  = str(row[col_sup]).strip()
        if not name or not sup:
            continue
        # אם הערך השתנה – נעדכן
        if store["name_map"].get(name) != sup:
            store["name_map"][name] = sup
            added += 1

    vk_save(store)
    return added

def build_vlookup_sheet(datasheet_df: pd.DataFrame) -> pd.DataFrame:
    """
    כל שורה כלל 2 → 'סכום חובה' = |סכום|.
    שורת סיכום 20001 בזכות = סה״כ חובה של השורות שיש להן 'מס' ספק'.
    שורות בלי 'מס' ספק' – יצבעו בכתום בשלב העיצוב.
    """
    store = vk_load()
    name_map   = {str(k): v for k, v in store.get("name_map", {}).items()}
    amount_map = {float(k): v for k, v in store.get("amount_map", {}).items()}

    col_match = pick_col(datasheet_df, MATCH_COLS) or datasheet_df.columns[0]
    col_bamt  = pick_col(datasheet_df, BANK_AMTS)
    col_det   = pick_col(datasheet_df, DETAILS)

    match = pd.to_numeric(datasheet_df[col_match], errors="coerce").fillna(0).astype(int)
    bamt  = to_num(datasheet_df[col_bamt]) if col_bamt else pd.Series([np.nan] * len(datasheet_df))
    det   = datasheet_df[col_det].astype(str).fillna("")

    vk = datasheet_df.loc[match == 2, [col_det, col_bamt]].rename(
        columns={col_det: "פרטים", col_bamt: "סכום"}
    ).copy()

    if vk.empty:
        return pd.DataFrame(columns=["פרטים", "סכום", "מס' ספק", "סכום חובה", "סכום זכות"])

    def pick_supplier(row):
        s = str(row["פרטים"])
        # קודם לפי טקסט
        for k, v in name_map.items():
            if k and k in s:
                return v
        # אחר כך לפי סכום מוחלט
        try:
            key = round(abs(float(row["סכום"])), 2)
            return amount_map.get(key, "")
        except Exception:
            return ""

    vk["מס' ספק"]   = vk.apply(pick_supplier, axis=1)
    vk["סכום חובה"] = vk["סכום"].apply(lambda x: abs(x) if pd.notna(x) else 0.0)
    vk["סכום זכות"] = 0.0

    total_hova_with_supplier = vk.loc[
        vk["מס' ספק"].astype(str).str.len() > 0, "סכום חובה"
    ].sum()

    if total_hova_with_supplier and total_hova_with_supplier != 0:
        vk = pd.concat(
            [
                vk,
                pd.DataFrame(
                    [
                        {
                            "פרטים": "סה\"כ זכות – עם מס' ספק",
                            "סכום": 0.0,
                            "מס' ספק": 20001,
                            "סכום חובה": 0.0,
                            "סכום זכות": round(float(total_hova_with_supplier), 2),
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )

    return vk

# ---------------- Rules 1–4 ----------------
def apply_rules_1_4(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    col_match = pick_col(out, MATCH_COLS) or out.columns[0]
    col_code  = pick_col(out, BANK_CODES)
    col_bamt  = pick_col(out, BANK_AMTS)
    col_aamt  = pick_col(out, BOOKS_AMTS)
    col_ref1  = pick_col(out, REF1S)
    col_ref2  = pick_col(out, REF2S)
    col_date  = pick_col(out, DATES)
    col_det   = pick_col(out, DETAILS)

    if col_match not in out.columns:
        out[col_match] = 0

    match = pd.to_numeric(out[col_match], errors="coerce").fillna(0).astype(int)
    code  = to_num(out[col_code]) if col_code else pd.Series([np.nan]*len(out))
    bamt  = to_num(out[col_bamt]) if col_bamt else pd.Series([np.nan]*len(out))
    aamt  = to_num(out[col_aamt]) if col_aamt else pd.Series([np.nan]*len(out))
    datev = norm_date(pd.to_datetime(out[col_date], errors="coerce")) if col_date else pd.Series([pd.NaT]*len(out))
    det   = out[col_det].astype(str).fillna("") if col_det else pd.Series([""]*len(out))
    ref1  = out[col_ref1].astype(str).fillna("") if col_ref1 else pd.Series([""]*len(out))
    ref2  = out[col_ref2].astype(str).fillna("") if col_ref2 else pd.Series([""]*len(out))

    # 1: OV/RC 1:1
    bank_keys, books_keys = {}, {}
    for i in range(len(out)):
        if match.iat[i] != 0:
            continue
        if (
            pd.notna(code.iat[i])
            and int(code.iat[i]) in OVRC_CODES
            and pd.notna(bamt.iat[i])
            and bamt.iat[i] < 0
            and pd.notna(datev.iat[i])
        ):
            k = (round(abs(float(bamt.iat[i])), 2), datev.iat[i])
            bank_keys.setdefault(k, []).append(i)

    for j in range(len(out)):
        if match.iat[j] != 0:
            continue
        if (
            pd.notna(aamt.iat[j])
            and aamt.iat[j] > 0
            and pd.notna(datev.iat[j])
            and str(ref1.iat[j]).upper().startswith(("OV", "RC"))
        ):
            k = (round(abs(float(aamt.iat[j])), 2), datev.iat[j])
            books_keys.setdefault(k, []).append(j)

    for k, bidx in bank_keys.items():
        if len(bidx) == 1 and len(books_keys.get(k, [])) == 1:
            i = bidx[0]
            j = books_keys[k][0]
            if match.iat[i] == 0 and match.iat[j] == 0:
                match.iat[i] = 1
                match.iat[j] = 1

    # 2: Standing orders (סימון בלבד)
    for i in range(len(out)):
        if match.iat[i] == 0 and pd.notna(code.iat[i]) and int(code.iat[i]) in STANDING_CODES:
            match.iat[i] = 2

    # 3: יסומן בשלב process_workbook (תלוי עזר; ללא בדיקת תאריך)

    # 4: שיקים ספקים (Ref1 בנק ↔ Ref2 ספרים) + טולרנס
    bank_idx = [
        i for i in range(len(out))
        if match.iat[i] == 0
        and pd.notna(code.iat[i])
        and int(code.iat[i]) == RULE4_CODE
        and str(ref1.iat[i]).strip()
        and pd.notna(bamt.iat[i])
    ]
    books_idx = [
        j for j in range(len(out))
        if match.iat[j] == 0
        and str(ref1.iat[j]).upper().startswith("CH")
        and str(ref2.iat[j]).strip()
        and pd.notna(aamt.iat[j])
    ]
    used = set()
    for i in bank_idx:
        ref_b = only_digits(ref1.iat[i])
        ab = abs(float(bamt.iat[i]))
        for j in books_idx:
            if j in used or match.iat[j] != 0:
                continue
            if only_digits(ref2.iat[j]) != ref_b:
                continue
            aj = abs(float(aamt.iat[j]))
            if abs(aj - ab) <= RULE4_EPS:
                match.iat[i] = 4
                match.iat[j] = 4
                used.add(j)
                break

    out[col_match] = match
    return out

# ---------------- Rules 5–12 ----------------
def apply_rules_5_12(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    col_match = pick_col(out, MATCH_COLS) or out.columns[0]
    col_code  = pick_col(out, BANK_CODES)
    col_bamt  = pick_col(out, BANK_AMTS)
    col_det   = pick_col(out, DETAILS)

    if col_match not in out.columns:
        out[col_match] = 0
    match = pd.to_numeric(out[col_match], errors="coerce").fillna(0).astype(int)
    code  = to_num(out[col_code]) if col_code else pd.Series([np.nan]*len(out))
    bamt  = to_num(out[col_bamt]) if col_bamt else pd.Series([np.nan]*len(out))
    det   = out[col_det].astype(str).fillna("") if col_det else pd.Series([""]*len(out))

    m5  = (match == 0) & (code.isin(list(RULE5_CODES))) & (bamt > 0) & (bamt <= 1000)
    match.loc[m5] = 5

    m6  = (match == 0) & (code == 175) & (bamt < 0) & (det == RULE6_COMPANY)
    match.loc[m6] = 6

    m7  = (match == 0) & (code == RULE7_CODE) & (bamt < 0) & (det == RULE7_PHRASE)
    match.loc[m7] = 7

    m8  = (match == 0) & (code == RULE8_CODE) & (bamt < 0) & (det == RULE8_PHRASE)
    match.loc[m8] = 8

    m9  = (match == 0) & (code == RULE9_CODE) & (bamt < 0) & (det == RULE9_PHRASE)
    match.loc[m9] = 9

    m10 = (match == 0) & (code.isin(list(RULE10_CODES))) & (bamt.notna()) & (bamt != 0)
    match.loc[m10] = 10

    match = rule11_placeholder(out.assign(**{col_match: match}), col_match, pick_col(out, BANK_CODES),
                               pick_col(out, BANK_AMTS), pick_col(out, DETAILS))
    match = rule12_placeholder(out.assign(**{col_match: match}), col_match, pick_col(out, BANK_CODES),
                               pick_col(out, BANK_AMTS), pick_col(out, DETAILS))

    out[col_match] = match
    return out

# ---------------- Styling & print ----------------
def style_and_print(wb):
    for ws in wb.worksheets:
        ws.sheet_view.rightToLeft = True
        ws.page_setup.paperSize = 9      # A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)

    # צביעה כתומה לשורות ללא 'מס' ספק' בגיליון הוראת קבע
    if "הוראת קבע ספקים" in wb.sheetnames:
        ws = wb["הוראת קבע ספקים"]
        header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        col_supplier = header.get("מס' ספק")
        if col_supplier:
            orange = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=col_supplier).value
                if v in ("", None):
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = orange
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

# ---------------- Processing ----------------
def process_workbook(main_bytes: bytes, aux_bytes: bytes | None):
    # קריאה
    wb = load_workbook(io.BytesIO(main_bytes), data_only=True)
    ws = wb["DataSheet"] if "DataSheet" in wb.sheetnames else wb.worksheets[0]
    df = ws_to_df(ws)
    if df.empty:
        return None, None, None

    # 1–4
    df = apply_rules_1_4(df)

    # === כלל 3 – סכומים זהים בלבד (ללא דרישת תאריך) ===
    st.session_state["_rule3_mismatches_df"] = None
    if aux_bytes is not None:
        aux_wb = load_workbook(io.BytesIO(aux_bytes), data_only=True)
        a_ws   = aux_wb.worksheets[0]
        a_df   = ws_to_df(a_ws)

        c_dt   = pick_col(a_df, AUX_DATE_KEYS)   # תאריך/חותמת אירוע
        c_amt  = pick_col(a_df, AUX_AMT_KEYS)    # אחרי ניכוי
        c_pay  = pick_col(a_df, AUX_PAYNO_KEYS)  # מס' תשלום

        if c_dt and c_amt:
            a_dt  = pd.to_datetime(a_df[c_dt], errors="coerce")               # אירוע
            a_amt = pd.to_numeric(a_df[c_amt], errors="coerce").round(2)
            groups = (pd.DataFrame({"evt": a_dt, "amt": a_amt})
                        .dropna(subset=["evt"])
                        .groupby("evt")["amt"].sum().round(2).to_dict())

            pays_by_evt = {}
            if c_pay:
                pays_by_evt = (pd.DataFrame({"evt": a_dt, "pay": a_df[c_pay].astype(str).str.strip()})
                                 .groupby("evt")["pay"]
                                 .apply(lambda s: set(s.dropna().astype(str)))
                                 .to_dict())

            col_match = pick_col(df, MATCH_COLS) or df.columns[0]
            col_code  = pick_col(df, BANK_CODES)
            col_bamt  = pick_col(df, BANK_AMTS)
            col_det   = pick_col(df, DETAILS)
            col_ref1  = pick_col(df, REF1S)
            col_aamt  = pick_col(df, BOOKS_AMTS)

            match = pd.to_numeric(df[col_match], errors="coerce").fillna(0).astype(int)
            code  = to_num(df[col_code]) if col_code else pd.Series([np.nan]*len(df))
            bamt  = to_num(df[col_bamt]).round(2) if col_bamt else pd.Series([np.nan]*len(df))
            det   = df[col_det].astype(str).fillna("") if col_det else pd.Series([""]*len(df))
            ref1  = df[col_ref1].astype(str).str.strip() if col_ref1 else pd.Series([""]*len(df))
            aamt  = to_num(df[col_aamt]).round(2) if col_aamt else pd.Series([np.nan]*len(df))

            bank_mask = (match == 0) & (code == TRANSFER_CODE) & (bamt > 0) & (det.str.contains(TRANSFER_PHRASE, na=False))
            mismatches = []

            for evt, evt_sum in groups.items():
                # ספרים לפי payset של האירוע
                payset = pays_by_evt.get(evt, set())
                books_idx = []
                books_sum = 0.0
                if payset is not None and len(payset) > 0 and col_ref1 and col_aamt:
                    books_idx = df.index[(match == 0) & (ref1.astype(str).isin(payset))].tolist()
                    if books_idx:
                        books_sum = float(pd.to_numeric(aamt.iloc[books_idx], errors="coerce").fillna(0).sum().round(2))

                # בנק – כל השורות שסכומן |bamt| == |evt_sum|
                bank_idx = df.index[bank_mask & (bamt.abs().sub(abs(evt_sum)).abs() <= RULE3_AMOUNT_EPS)].tolist()

                if bank_idx and books_idx:
                    # התאמה חייבת להיות שוויון בערך מוחלט
                    if abs(abs(books_sum) - abs(evt_sum)) <= RULE3_AMOUNT_EPS:
                        for i in bank_idx:
                            if match.iat[i] in (0, 2):
                                match.iat[i] = 3
                        for j in books_idx:
                            if match.iat[j] in (0, 2):
                                match.iat[j] = 3
                    else:
                        mismatches.append({
                            "אירוע": str(evt),
                            "סכום בעזר (אחרי ניכוי)": float(evt_sum),
                            "סכום בספרים (סיכום)": float(books_sum),
                            "פער |ספרים|-|עזר|": float(round(abs(abs(books_sum) - abs(evt_sum)), 2)),
                            "count_בנק": len(bank_idx),
                            "count_ספרים": len(books_idx),
                        })
                else:
                    mismatches.append({
                        "אירוע": str(evt),
                        "סכום בעזר (אחרי ניכוי)": float(evt_sum),
                        "סכום בספרים (סיכום)": float(books_sum) if books_idx else np.nan,
                        "פער |ספרים|-|עזר|": np.nan,
                        "count_בנק": len(bank_idx),
                        "count_ספרים": len(books_idx),
                    })

            df[col_match] = match
            if mismatches:
                st.session_state["_rule3_mismatches_df"] = pd.DataFrame(mismatches)

    # 5–12 (רק על 0)
    df = apply_rules_5_12(df)

    # גיליון הוראת קבע ספקים
    vk_df = build_vlookup_sheet(df)

    # יצוא עם עיצוב + גיליון בקרה לכלל 3 (אם יש)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as wr:
        df.to_excel(wr, index=False, sheet_name="DataSheet")
        counts = pd.to_numeric(
            df[pick_col(df, MATCH_COLS) or df.columns[0]],
            errors="coerce"
        ).fillna(0).astype(int).value_counts().sort_index()
        pd.DataFrame({"מס": counts.index, "כמות": counts.values}).to_excel(
            wr, index=False, sheet_name="סיכום"
        )
        vk_df.to_excel(wr, index=False, sheet_name="הוראת קבע ספקים")

        misdf = st.session_state.get("_rule3_mismatches_df", None)
        if misdf is not None and not misdf.empty:
            misdf.to_excel(wr, index=False, sheet_name="פערי סכומים – כלל 3")

    wb_out = load_workbook(io.BytesIO(buffer.getvalue()))
    style_and_print(wb_out)
    final = io.BytesIO()
    wb_out.save(final)
    return df, vk_df, final.getvalue()

# ---------------- UI ----------------
c1, c2 = st.columns([2, 2])
main_file = c1.file_uploader("בחרי קובץ מקור – DataSheet בלבד", type=["xlsx"])
aux_file  = c2.file_uploader("⬆️ קובץ עזר להעברות (לכלל 3)", type=["xlsx"])
st.caption("VLOOKUP שומר מפות ב-rules_store.json (שם/סכום → מס' ספק).")

if st.button("הרצה 1–12"):
    if not main_file:
        st.error("נא להעלות קובץ מקור.")
    else:
        with st.spinner("מעבד..."):
            df_out, vk_out, out_bytes = process_workbook(
                main_file.read(), aux_file.read() if aux_file else None
            )
        if df_out is None:
            st.error("לא נמצאו נתונים.")
        else:
            st.success("מוכן!")
            col_match = pick_col(df_out, MATCH_COLS) or df_out.columns[0]
            cnt = pd.to_numeric(
                df_out[col_match], errors="coerce"
            ).fillna(0).astype(int).value_counts().sort_index()
            st.dataframe(
                pd.DataFrame({"מס": cnt.index, "כמות": cnt.values}),
                use_container_width=True,
            )
            st.download_button(
                "📥 הורד קובץ מעודכן",
                data=out_bytes,
                file_name="התאמות_1_עד_12.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

# ניהול מפות ל-VLOOKUP
st.divider()
st.subheader("🔎 VLOOKUP – הוראת קבע ספקים (עריכה, שמירה וייבוא מאקסל)")
store = vk_load()

with st.expander("מפות מיפוי (נשמר ל-rules_store.json)", expanded=False):
    # הוספה/עדכון לפי שם (contains)
    t1, t2 = st.columns([2, 1])
    nm = t1.text_input("מיפוי לפי 'פרטים' (contains)")
    sp = t2.text_input("מס' ספק")
    if st.button("➕ הוסף/עדכן לפי שם"):
        if nm and sp:
            store["name_map"][nm] = sp
            vk_save(store)
            st.success("נשמר לפי שם.")

    # הוספה/עדכון לפי סכום מוחלט
    t3, t4 = st.columns([1, 1])
    amt = t3.number_input("מיפוי לפי סכום (ערך מוחלט)", step=0.01, format="%.2f")
    sp2 = t4.text_input("מס' ספק", key="vk2")
    if st.button("➕ הוסף/עדכן לפי סכום"):
        try:
            store["amount_map"][str(round(abs(float(amt)), 2))] = sp2
            vk_save(store)
            st.success("נשמר לפי סכום.")
        except Exception as e:
            st.error(str(e))

    # ייבוא מאקסל – פרטים + מס' ספק
    st.markdown("---")
    up_map = st.file_uploader(
        "ייבוא קובץ מיפוי ספקים (עמודה 1 – פרטים, עמודה 2 – מס' ספק)",
        type=["xlsx"],
        key="vk_upload",
    )
    if up_map is not None and st.button("⬆️ ייבוא מיפוי מאקסל"):
        try:
            added = import_name_map_from_excel(up_map, store)
            st.success(f"התווספו/עודכנו {added} רשומות מהמיפוי.")
        except Exception as e:
            st.error(f"שגיאה בייבוא: {e}")
